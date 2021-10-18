import logging
import openpyxl


class Testdata:

    testdatavalue= [{"Username":"test@email.com", "Password":"abcabc", "Course_Name":"JavaScript", "Card_Number":"4242424242424242", "Expiry_Date":"0223", "CVV":"086", "Country":"Haiti"}]
    @staticmethod
    def dataexcelinput(testname):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Test Data\\TestExceltwo.xlsx")

        sheet=book.active

        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == testname:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value
        return [Dict]



